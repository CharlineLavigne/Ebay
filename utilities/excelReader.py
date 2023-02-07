import openpyxl

from openpyxl.styles import PatternFill


class ExcelReader:

    def read_excel_data(self, file, sheet_tested, col_tested):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheet_tested]
        excel_data_array = []

        for r in range(col_tested, col_tested + 1):
            for c in range(1, 3):
                excel_data_array.append(sheet.cell(r, c).value)
                print(sheet.cell(r, c).value)

        return excel_data_array

    def write_test_results(self, file, sheet_tested, row_result, cell_result, test_result):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheet_tested]
        sheet.cell(row_result, cell_result).value = test_result
        workbook.save(file)

    def fill_color(self, file, sheet_tested, row_result, cell_result, color):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheet_tested]
        color_filled = PatternFill(start_color=color, end_color=color, fill_type='solid')
        sheet.cell(row_result, cell_result).fill = color_filled
        workbook.save(file)



